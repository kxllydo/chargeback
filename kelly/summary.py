from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
import pandas as pd
import math
import time

def addDataAndHeader (wb, pyxlWS, path, columnNum, header, width = 0, dataList = []):
    '''
    This adds all of the headers to the group summary sheet and their respective values
    @param wb is the workbook loaded using openpyxl
    @param ws is the worksheet opened using openpyxl
    @param path is a string of the path to the excel workbook
    @param columnNum is the column you want to add the header to
    @param header is the header you want to add to the excel sheet
    @param width is the width of the column
    @param dataList is the list of the data you want in the column
    @return completed data and header
    '''
    cell = pyxlWS.cell(row = 1, column = columnNum)
    cell.value = header
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal = "center")

    for index, value in enumerate(dataList, start=2):
        cell = pyxlWS.cell(row=index, column=columnNum)
        cell.value = value
        if (len(dataList) != 0 ) and (isinstance(dataList[0], float)):
            cell.style = "Currency"

    if width != 0:
       pyxlWS.column_dimensions[chr(64 + columnNum)].width = width 

    wb.save(path)

def groupCostMerger(sheet):
    """
    Combines the same groups and sums their costs
    @param sheet is the sheet read in by using pandas
    @return a dictionary of the groups and their costs
    """
    colLength = len(sheet["Group"]) - 1
    applications = sheet.loc[0:colLength, "Group"]
    sumDict = {}
    uniqueApps = []

    
    for index, value in enumerate(applications, start = 0):
        cost = sheet.loc[index,"February (2024)"]
        if math.isnan(cost):
            cost = 0
        if value not in uniqueApps:
            uniqueApps.append(value)
            sumDict[value] = cost
        else:
            sumDict[value] += cost

    return sumDict

def merger(sheet, header1, header2 = "Group"):
    """
    General merger to assign relationship between non cost categories
    @param sheet is the sheet read in by using pandas
    @param header1 is the header of the unique data you want to access
    @param header2 is either the Group or Applications
    @return a dictionary of the group and other dataset
    """
    colLength = len(sheet[header2]) - 1
    applications = sheet.loc[0:colLength, header2]
    dict = {}
    uniqueApps = []

    if header2 == "Group":
        uniqueApps = []
        for index, value in enumerate(applications, start = 0):
            if value not in uniqueApps:
                dict[value] = sheet.loc[index, header1]
                uniqueApps.append(value)
    else:
        for index, value in enumerate(applications, start = 0):
            dict[value] = sheet.loc[index, header1]

    return dict


def infracharge(total):
    """
    Finds the value of the infracharge and makes a list of it
    @param sumSheet is the summary sheet opened with pandas
    @param total is the dictionary of all the groups and the total cost
    """
    groupNum = len(total) - 1
    print(groupNum)
    cost = round((total["Infrastructure"]/groupNum), 2)
    length = len(total.values())
    data  = []
    for i in range(length):
        data.append(cost)
    return data

def creategroupSummarySheet(wb, sumSheet, path):
    """
    Creates the group summary sheet and fills it with cost, PC, and AC
    @param wb is the workbook loaded using openpyxl
    @param sumSheet is the summary worksheet opened with pandas
    @param path is a string path to the workbook
    """
    pc = merger(sumSheet, "PC")
    ac = merger(sumSheet, "AC")
    cost = groupCostMerger(sumSheet)
    charge = infracharge(cost)

    wb.create_sheet("Group Summary")
    ws = wb["Group Summary"]
    headers = ["Applications", "Amount", "Infra Charge", "Adjustments", "Total Charge",	"Profit Center", "Account Code"]
    
    addDataAndHeader(wb, ws, path, 1, headers[0], 49, list(cost))
    addDataAndHeader(wb, ws, path, 2, headers[1], 14.55, list(cost.values()))
    addDataAndHeader(wb, ws, path, 3, headers[2], 12.64, charge)
    addDataAndHeader(wb, ws, path, 4, headers[3], 13.91)
    addDataAndHeader(wb, ws, path, 5, headers[4], 16.64)
    addDataAndHeader(wb, ws, path, 5, headers[5], 15.45, list(pc.values()))
    addDataAndHeader(wb, ws, path, 6, headers[6], 15.45, list(ac.values()))

def createCustomerSheet(wb, sumSheet, path):
    """
    Creates customer chargeback sheet for the month
    @param wb is the workbook opened by openpyxl
    @param sumsheet is the summary sheet opened by pandas
    @param path is the path to the excel workbook
    """
    pc = merger(sumSheet, "PC")
    ac = merger(sumSheet, "AC")
    cost = groupCostMerger(sumSheet)
    charge = infracharge(cost)
    owner = merger(sumSheet, "Owner")

    for key in cost:
        cost[key] += charge[0]

    pc.pop("Infrastructure")        #find a way to shorten this
    ac.pop("Infrastructure")
    owner.pop("Infrastructure")
    cost.pop("Infrastructure")

    wb.create_sheet("Customer Chargeback")

    ws = wb["Customer Chargeback"]
    headers = ["Owner","Applications","February (2024)","Profit Center", "AC"]
    addDataAndHeader(wb, ws, path, 1, headers[0], 31.64, list(owner.values()))
    addDataAndHeader(wb, ws, path, 2, headers[1], 32.64,  list(owner))
    addDataAndHeader(wb, ws, path, 3, headers[2], 19, list(cost.values()))
    addDataAndHeader(wb, ws, path, 4, headers[3], 15.91, list(pc.values()))
    addDataAndHeader(wb, ws, path, 5, headers[4], 16.64, list(ac.values()))


def createTaxSheet(wb, path, tax):
    """
    Creates tax chargeback sheet for the month
    @param wb is the workbook opened by openpyxl
    @param path is a string of the path to the excel workbook
    @param tax is a float of the tax percentage
    """
    customerSheet = wb["Customer Chargeback"]
    copy = wb.copy_worksheet(customerSheet)
    copy.title = "Tax Chargeback"
    wb.save(path)

    pandaWs = pd.read_excel(path, sheet_name="Tax Chargeback")
    cost = merger(pandaWs, "February (2024)", "Applications")
    ws = wb["Tax Chargeback"]

    salesTax = []
    taxPercent = []

    for i in range(len(cost)):
        taxPercent.append(tax)
    for key in cost:
        taxes = tax * cost[key]
        salesTax.append(taxes) 
        cost[key] += taxes

    addDataAndHeader(wb, ws, path, 6, "Sales tax %", 13.45, taxPercent)
    addDataAndHeader(wb, ws, path, 7, "Sales Tax", 13.55, salesTax)
    addDataAndHeader(wb, ws, path, 8, "Total", 16.82, list(cost.values()))
    
    wb.save(path)

def createChargebacks(wb, path, sumSheet, tax):
    """
    Creates tax and customer chargeback sheet for the current month
    @param wb is the worbook opened by openpyxl
    @param path is a string of the path to the workbook
    @param sumSheet is the summary sheet opened by pandas
    @param tax is a float of the sales tax percentage
    """

    createCustomerSheet(wb, sumSheet, path)
    createTaxSheet(wb, path, 0.0305)

    pyxlWS = wb["Customer Chargeback"]
    finalCosts = merger(pd.read_excel(path, "Tax Chargeback"), "Total", "Applications")
    addDataAndHeader(wb, pyxlWS, path, 3, "February (2024)", 19,  list(finalCosts.values()))