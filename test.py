from openpyxl import Workbook, load_workbook
import pandas as pd
import csv
import json
import math

def addDataAndHeader (wb, ws, path, dataList, columnNum, header):
    '''
    This adds all of the headers to the group summary sheet
    @param wb is the workbook loaded using openpyxl
    @param ws is the worksheet opened using openpyxl
    @path is a string of the path to the excel workbook
    @columnNum is the column you want to add the header to
    @header is the header you want to add to the excel sheet
    '''
    cell = ws.cell(row = 1, column = columnNum)
    cell.value = header
    for index, value in enumerate(dataList, start=2):
        cell = ws.cell(row=index, column=columnNum)
        cell.value = value
    wb.save(path)

def rgComparer(path):
    """
    Tells you which resource groups were added or deleted from the previous month
    @param path is the path to the excel workbook
    """
    comparison = pd.read_excel(path, sheet_name = "RG Comparison")
    lastMonthRg = comparison["Last Month RG"].tolist()
    currentMonthRg = comparison["Current Month RG"].tolist()

    deleted, added = []
    for rgs in lastMonthRg:
        if lastMonthRg not in currentMonthRg:
            deleted.append(rgs)

    for rg in currentMonthRg:
        if currentMonthRg not in lastMonthRg:
            added.append(rg)
    
    wb = load_workbook(path)
    sheet = wb["RG Comparison"]
    addDataAndHeader(wb, sheet, path, deleted, 4, "Deleted")
    addDataAndHeader(wb, sheet, path, added, 5, "Added")

def groupCostMerger(sheet):
    """
    Combines the same groups and sums their costs
    @param sheet is the sheet read in by using pandas
    @return a dictionary of the groups and their costs
    """
    colLength = len(sheet["Group"]) - 1
    applications = sheet.loc[0:colLength, "Group"]
    sumDict = {}
    uniqueApps = []

    for index, value in enumerate(applications, start = 0):
        cost = sheet.loc[index,"February (2024)"]
        if math.isnan(cost):
            cost = 0
        if value not in uniqueApps:
            uniqueApps.append(value)
            sumDict[value] = cost
        else:
            sumDict[value] += cost
    return sumDict

def merger(sheet, header):
    """
    General merger to assign relationship between non cost categories
    @param sheet is the sheet read in by using pandas
    @header is the header of the data you want to access
    @return a dictionary of the group and other dataset
    """
    colLength = len(sheet["Group"]) - 1
    applications = sheet.loc[0:colLength, "Group"]
    dict = {}
    uniqueApps = []

    for index, value in enumerate(applications, start = 0):
        if value not in uniqueApps:
            dict[value] = sheet.loc[index, header]
            uniqueApps.append(value)
    return dict

def creategroupSummarySheet(wb, path, groups, profit, allocation):
    """
    Creates the group summary sheet and fills it with cost, PC, and AC
    @param wb is the workbook loaded using openpyxl
    @param path is a string path to the workbook
    @param groups is the dicitonary of group and cost
    @param profit is the dictionary of group and PC
    @param allocaiton is the dictionary of group and AC
    """
    wb.create_sheet("Group Summary")
    ws = wb["Group Summary"]
    headers = ["Applications", "Amount", "Infra Charge", "Adjustments", "Total Charge",	"Profit Center", "Account Code"]
    for i, header in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=header)
    wb.save(path)
    fillColumn(wb, ws, path, 1, list(groups))
    fillColumn(wb, ws, path, 2, list(groups.values()))
    fillColumn(wb, ws, path, 6, list(profit.values()))
    fillColumn(wb, ws, path, 7, list(allocation.values()))

def fillColumn(wb, ws, path, colNum, data):
    row = 2
    for d in data:
        ws.cell(row=row, column = colNum).value = d
        row+=1
    wb.save(path)



if __name__ == "__main__":
    excel = "c:\\Users\\do-kelly\\Downloads\\help.xlsx"
    sheet = "Summary"

    wb = load_workbook(excel)
    ws = pd.read_excel(excel, sheet_name=sheet)

    pc = merger(ws, "PC")
    ac = merger(ws, "AC")
    cost = groupCostMerger(ws)

    creategroupSummarySheet(wb, excel, cost, pc, ac)
