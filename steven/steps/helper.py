import os
from openpyxl.styles import Font, Alignment
from openpyxl.utils.cell import get_column_letter

BLANK_VALUE = "(blank)"     # what to replace n/a values with
ROW_PADDING = 0             # how many rows is the header below row 1?

# For RG Comparison Sheet
RGSHEET = "RG Comparison"
LAST_MONTH_RGS = "Last Month RG"
CURRENT_MONTH_RGS = "Current Month RG"

# For Summary Sheet
SUMMARY = "Summary"

def check_for_perms(path = "."):
    """
        Checks for exist, read, and write perms.
        Prints error messages for lacking perms.
        Terminates program if lacking any perm.

        param @path (string)    : Path of file to check
    """
    if not os.path.exists(path):
        print(f"{os.path.abspath(path)} does not exist.")
        exit(-1)
    if not os.access(path, os.R_OK):
        print(f"Please grant READ permissions to {os.path.abspath(path)}")
        exit(-1)
    if not os.access(path, os.W_OK):
        print(f"Please grant WRITE permissions to {os.path.abspath(path)}")
        exit(-1)

def addColumn(worksheet, columnIndex, header, width = 0, data = [], format = {}):
    worksheet.column_dimensions[get_column_letter(columnIndex)].width = width

    cell = worksheet.cell(row = 1 + ROW_PADDING, column = columnIndex)
    cell.value = header
    cell.font = Font(bold = True)
    cell.alignment = Alignment(horizontal = "center")

    for index, value in enumerate(data, start = 2 + ROW_PADDING):
        cell = worksheet.cell(row = index, column = columnIndex)
        cell.value = value

        for key, value in format.items():
            cell[key] = value

def addRow(worksheet, rowIndex = -1, data = []):
    if rowIndex == -1:
        rowIndex = worksheet.max_row + 1
    worksheet.insert_rows(rowIndex, 1)

    for index, value in enumerate(data, start = 1):
        worksheet.cell(row = rowIndex, column = index, value = value)