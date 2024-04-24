import openpyxl as excel
import pandas as pd

def step_one(billing_file, excel_file):
    workbook = excel.load_workbook(excel_file)
    if "RG Comparison" not in workbook.sheetnames:
        workbook.create_sheet("RG Comparison")
    rg_cost_sheet = workbook["RG Comparison"]

    data = pd.read_csv(billing_file)
    data = data.filter(items = ["ResourceGroup", "Cost"])
    data.fillna("(blank)", inplace = True)
    data["RG_Lower"] = data["ResourceGroup"].str.lower()
    data["Current Month RG"] = data["ResourceGroup"]
    data = data.groupby("RG_Lower", as_index = False).agg(
        {"Current Month RG": "first",
         "Cost": "sum"})

    # Adding Headers to the Worksheet
    rg_cost_sheet.column_dimensions["A"].width = 50.45
    rg_cost_sheet.column_dimensions["B"].width = 50.45
    rg_cost_sheet.column_dimensions["C"].width = 11.18

    rg_cost_sheet.cell(row = 1, column = 1, value = "Last Month RG")
    rg_cost_sheet.cell(row = 1, column = 2, value = "Current Month RG")
    rg_cost_sheet.cell(row = 1, column = 3, value = "Cost")
    
    for col in range(1, rg_cost_sheet.max_column + 1):
        cell = rg_cost_sheet.cell(row = 1, column = col)
        cell.font = excel.styles.Font(bold = True)
        cell.alignment = excel.styles.Alignment(horizontal = "center")

    # Filling out data
    for i, row in data.iterrows():
        rg_cost_sheet.cell(row = i + 2, column = 2, value = row["Current Month RG"])
        rg_cost_sheet.cell(row = i + 2, column = 3, value = row["Cost"])

    workbook.save(excel_file)
    workbook.close()