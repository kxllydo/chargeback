import openpyxl as excel
from openpyxl.utils.cell import get_column_letter as get_letter

def step_three(excel_file, current_month):

    workbook = excel.load_workbook(excel_file)
    if "Summary" not in workbook.sheetnames:
        print("Summary sheet does not exist in chargeback")
        exit(-1)

    rg_cost_sheet = workbook["RG Comparison"]
    summary_sheet = workbook["Summary"]

    rgs_in_summary = [""]
    for row, rg in enumerate(list(summary_sheet.columns)[-2]):
        if row < 1: continue
        if not rg.value: break
        rgs_in_summary.insert(row + 1, rg.value.lower())

    # Header for Month (Year)
    header_pos = summary_sheet.max_column - 2
    summary_sheet.insert_cols(header_pos)
    header = summary_sheet.cell(row = 1, column = header_pos)
    header.value = current_month
    header.font = excel.styles.Font(bold = True)

    # Fill in cost data
    new_rgs = []
    for row, rg in enumerate(list(rg_cost_sheet.columns)[1]):
        try:
            index = rgs_in_summary.index(rg.value.lower())
            cell = summary_sheet.cell(row = index + 1,
                                       column = header_pos)
            cell.value = rg_cost_sheet.cell(row = row + 1, column = 3).value
            cell.number_format = excel.styles.numbers.BUILTIN_FORMATS[44]
        except ValueError:
            new_rgs.append(rg.value)

    # print(new_rgs)

    # Adjust widths
    summary_sheet.column_dimensions[get_letter(header_pos)].width = 17.78
    summary_sheet.column_dimensions[get_letter(header_pos + 1)].width = 17.78
    summary_sheet.column_dimensions[get_letter(header_pos + 2)].width = 58.44
    summary_sheet.column_dimensions[get_letter(header_pos + 3)].width = 58.44

    workbook.save(excel_file)
    workbook.close()