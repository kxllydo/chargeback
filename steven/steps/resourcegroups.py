import openpyxl as excel
import pandas as pd

from . import helper

def extract_billing(billing):
    """
        Extracts the CSV file into a Pandas DataFrame.
        Replaces all blank values with helper.BLANK_VALUE.
        Processes DF PivotTable-like, with columns Resource Group & Cost

        In procesing, when resource group is helper.BLANK_VALUE, AND when
        * ConsumedService = microsoft.visualstudio: resource group is taken from ResourceId, where the last segment is assumed to be the organizatio name (.../organizations/name)
        * ConsumedService = helper.BLANK_VALUE: resource group remains helper.BLANK_VALUE

        @param billing (string): path to billing CSV
        @returns DataFrame obj, with resource group and cost
    """
    helper.check_for_perms(billing)

    data = pd.read_csv(billing)
    data = data.filter(items = ["ConsumedService", "ResourceId", "ResourceGroup", "Cost"]).fillna(helper.BLANK_VALUE)

    # Case #1: ConsumedService = microsoft.visualstudio, ResourceGroup = helper.BLANK_VALUE
    data.loc[(data["ConsumedService"] == "microsoft.visualstudio") & (data["ResourceGroup"] == helper.BLANK_VALUE), "ResourceGroup"] = data["ResourceId"].str.split("/").str[-1]
    
    # Logging for more helper.BLANK_VALUE resource group cases
    errlog = data.loc[(data["ResourceGroup"] == helper.BLANK_VALUE) & (data["ConsumedService"] != helper.BLANK_VALUE)]
    if not errlog.empty: print(errlog.to_string(max_rows = None))

    data["RG_Lower"] = data["ResourceGroup"].str.lower()
    data[helper.CURRENT_MONTH_RGS] = data["ResourceGroup"]
    data = data.groupby("RG_Lower", as_index = False).agg(
        {helper.CURRENT_MONTH_RGS: "first", 
         "Cost": "sum"})
    
    return data

def step_one(billing, chargeback):
    helper.check_for_perms(billing)
    helper.check_for_perms(chargeback)

    data = extract_billing(billing)

    workbook = excel.load_workbook(chargeback)
    worksheet = workbook[helper.RGSHEET] if helper.RGSHEET in workbook.sheetnames else workbook.create_sheet(helper.RGSHEET)
    worksheet.delete_cols(2, 4)

    helper.addColumn(worksheet, 1, helper.LAST_MONTH_RGS, 50.45)
    helper.addColumn(worksheet, 2, helper.CURRENT_MONTH_RGS, 50.45, data[helper.CURRENT_MONTH_RGS])
    helper.addColumn(worksheet, 3, "Cost", 11.18, data["Cost"])
    
    grand_total = sum(data["Cost"].to_list())
    helper.addRow(worksheet, data = ["", "", grand_total])

    workbook.save(chargeback)
    workbook.close()

def step_two(chargeback1, chargeback2):
    helper.check_for_perms(chargeback1)
    helper.check_for_perms(chargeback2)

    data = pd.read_excel(chargeback1, sheet_name = helper.RGSHEET)
