import openpyxl as excel
import pandas as pd
from openpyxl.utils.cell import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from . import helper as helper

def extract_data(chargeback):
    helper.check_for_perms(chargeback)

    workbook = excel.load_workbook(chargeback)
    if helper.SUMMARY not in workbook.sheetnames:
        print(f"{helper.SUMMARY} not found in {chargeback}.")
        exit(-1)

    summary_data = pd.read_excel(chargeback, sheet_name = helper.SUMMARY, header = helper.ROW_PADDING)
    summary_data = summary_data["Resource Group"].str.lower().to_list()
    summary_data_copy = summary_data.copy()
    for _ in range(helper.ROW_PADDING + 2): summary_data.insert(0, "")

    rgs_data = pd.read_excel(chargeback, sheet_name = helper.RGSHEET, header = helper.ROW_PADDING)
    found_in_summary = []
    not_found_in_summary = []
    for _, row in rgs_data.iterrows():
        try:
            index = summary_data.index(row[helper.CURRENT_MONTH_RGS].lower()) - helper.ROW_PADDING - 2
            summary_data_copy.remove(row[helper.CURRENT_MONTH_RGS].lower())

            for _ in range(index - len(found_in_summary) + 1): found_in_summary.append("")
            found_in_summary[index] = row["Cost"]
        except ValueError:
            not_found_in_summary.append(row[helper.CURRENT_MONTH_RGS])

    print("These resource groups were not found in the summary sheet.")
    for v in not_found_in_summary: print(v)
    print("\nThese resource groups were foudn in the summary sheet, but are not in this month's list of resouce groups.")
    for v in summary_data_copy: print(v)

    workbook.close()
    return found_in_summary

def fix_formatting(sheet):
    new_colidx = sheet.max_column - 3
    sheet.column_dimensions[get_column_letter(new_colidx)].width = 17.78
    sheet.column_dimensions[get_column_letter(new_colidx + 1)].width = 17.78
    sheet.column_dimensions[get_column_letter(new_colidx + 2)].width = 58.44
    sheet.column_dimensions[get_column_letter(new_colidx + 3)].width = 58.44

def update_variance(chargeback, sheet):
    helper.check_for_perms(chargeback)

    data = pd.read_excel(chargeback, sheet_name = helper.SUMMARY)
    previous_month = data.columns[-5:-2][0]
    current_month = data.columns[-5:-2][1]
    data["Variance"] = data[previous_month] - data[current_month]
    
    helper.addColumn(sheet, -3, header = "Variance", width = 17.78, data = data["Variance"], format = {
        "number_format": excel.styles.numbers.BUILTIN_FORMATS[44]
    }, overwrite = True)

def step_three(chargeback, header):
    helper.check_for_perms(chargeback)

    data = extract_data(chargeback)
    workbook = excel.load_workbook(chargeback)
    summary_sheet = workbook[helper.SUMMARY]

    helper.addColumn(summary_sheet, -3, header = header, data = data, format = {
        "number_format": excel.styles.numbers.BUILTIN_FORMATS[44]
    }, overwrite = False)

    fix_formatting(summary_sheet)
    workbook.save(chargeback)

    update_variance(chargeback, summary_sheet)
    workbook.save(chargeback)
    workbook.close()

def step_four(chargeback1, chargeback2):
    helper.check_for_perms(chargeback1)
    helper.check_for_perms(chargeback2)

    data = pd.read_excel(chargeback1, sheet_name = helper.SUMMARY, header = helper.ROW_PADDING)

    workbook = excel.load_workbook(chargeback2)
    worksheet = workbook.create_sheet(helper.SUMMARY)

    #worksheet.append(data.columns.to_list())
    for r in dataframe_to_rows(data, index = False, header = True):
        worksheet.append(r)

    workbook.save(chargeback2)
    workbook.close()