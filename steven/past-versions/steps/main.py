import sys
import os
import datetime
import openpyxl as excel
import pandas as pd

#from steps.stepone import step_one
from stepthree import step_three

def check_for_perms(path):
    """
        Checks for exist, read, and write perms.
        Prints error messages for lacking perms.
        Terminates program if lacking any perm.

        param @path (string)    : Path of file to check
    """
    if not os.path.exists(path):
        print(f"{os.path.abspath(path)} does not exist.")
        exit(-1)
    if not os.access(path, os.R_OK):
        print(f"Please grant READ permissions to {os.path.abspath(path)}")
        exit(-1)
    if not os.access(path, os.W_OK):
        print(f"Please grant WRITE permissions to {os.path.abspath(path)}")
        exit(-1)

def step_one(billing, chargeback):
    """
        Compiles billing CSV into a makeshift pivot table, of resource groups and sum of costs for each resource group
        Creates a sheet inside chargeback.xlsx "RG Comparison"
        
        Edge Cases (when resource group is blank):
        * ComsumedService = microsoft.visualstudio
        * ConsumedService = blank
        (add more as it appears)

        @param billing (string)     : path to billing csv
        @param chargeback (string)  : path to chargeback xlsx
    """
    check_for_perms(billing)
    check_for_perms(chargeback)

    data = pd.read_csv(billing)
    data = data.filter(items = ["ConsumedService", "ResourceId", "ResourceGroup", "Cost"])
    data.fillna("(blank)", inplace = True)

    # Edge Case #1: ConsumedService = microsoft.visualstudio
    # Assigns resource group to organization group, found in column ResourceId (AB)
    # Under the assumption that ResourceId is formatted: .../organizations/organization-name
    data.loc[(data["ConsumedService"] == "microsoft.visualstudio") & (data["ResourceGroup"] == "(blank)"), "ResourceGroup"] = data["ResourceId"].str.split("/").str[-1]

    # Logging for any more edge cases, if they still exist
    # Will print for any row with a blank ResourceGroup.
    # Will also display ComputeService
    errlog = data.loc[data["ResourceGroup"] == "(blank)"]
    errlog = errlog.groupby("ConsumedService", as_index = False)
    errlog.agg({"ConsumedService" : "first", "ResourceGroup" : "count"})
    for _, v in errlog:
        print(v.to_string(max_rows = None))


    data["RG_Lower"] = data["ResourceGroup"].str.lower()
    data.rename(columns = {"ResourceGroup" : "Current Month RG"}, inplace = True)
    data = data.groupby("RG_Lower", as_index = False).agg(
        {"Current Month RG" : "first",
         "Cost" : "sum"})
    
    # Adding data to chargeback sheet
    workbook = excel.load_workbook(chargeback)
    if "RG Comparison" not in workbook.sheetnames:
        workbook.create_sheet("RG Comparison")
    rg_cost_sheet = workbook["RG Comparison"]

    rg_cost_sheet.column_dimensions["A"].width = 50.45
    rg_cost_sheet.column_dimensions["B"].width = 50.45
    rg_cost_sheet.column_dimensions["C"].width = 11.18
    rg_cost_sheet.cell(row = 1, column = 1, value = "Last Month RG")
    rg_cost_sheet.cell(row = 1, column = 2, value = "Current Month RG")
    rg_cost_sheet.cell(row = 1, column = 3, value = "Cost")
    for col in range(1, rg_cost_sheet.max_column + 1):
        cell = rg_cost_sheet.cell(row = 1, column = col)
        cell.font = excel.styles.Font(bold = True)
        cell.alignment = excel.styles.Alignment(horizontal = "center")

    for i, row in data.iterrows():
        rg_cost_sheet.cell(row = i + 2, column = 2, value = row["Current Month RG"])
        rg_cost_sheet.cell(row = i + 2, column = 3, value = row["Cost"])
    
    workbook.save(excel_file)
    workbook.close()

def step_four(billing, chargeback):
    """
        For dict of apps & regions
    """
    rg_regions_df = pd.read_csv(billing)
    app_rgs_df = pd.read_excel(chargeback, sheet_name = "Summary")

    rg_regions_df = rg_regions_df.filter(items = ["ResourceGroup", "ResourceLocation"])
    app_rgs_df = app_rgs_df.filter(items = ["Resource Group", "Application"])
    app_rgs_df.rename(columns = {"Resource Group" : "ResourceGroup"}, inplace = True) 
    
    result = pd.merge(rg_regions_df, app_rgs_df, on = "ResourceGroup")
    print(result)

if __name__ == "__main__":
    #current_date   = datetime.datetime.now()
    #month          = current_date.strftime("%B")
    #year           = current_date.strftime("%Y")

    # DELETE LATER
    month = "February"
    year = "2024"
    # DELETE LATER

    archive_directory = "../../archives"  # Path to top level parent folder of all files
    check_for_perms(archive_directory)
    check_for_perms(archive_directory + os.sep + year)
    check_for_perms(archive_directory + os.sep + year + os.sep + month)

    archive_directory   = archive_directory + os.sep + year + os.sep + month + os.sep
    billing_file        = archive_directory + "billing.csv"
    excel_file          = archive_directory + "chargeback.xlsx"
    check_for_perms(billing_file)
    check_for_perms(excel_file)

    step_one(billing_file, excel_file) # Completes current month's RG Comparison sheet, filling Current Month RG and Cost
                                       # Completes next months' RG Comparison sheet, fliling Last Month RG
    step_three(excel_file, 
               f"{month} ({year})")    # Adds current month's RGs to Summary Sheet