from kelly import summary
from kelly import rg
from steven import resourcegroups
from steven import summary as summary2
from openpyxl import load_workbook
import pandas as pd

import openpyxl as excel

if __name__ == "__main__":
    # Set CSV path
    billing = "../archives/2024/February/billing.csv"
    # Set XLSX path
    chargeback = "../archives/2024/February/chargeback.xlsx"
    chargeback2 = chargeback.replace(".xlsx", "2.xlsx")
    excel.Workbook().save(chargeback2)
    # Set Billing Cycle: (Example: February (2024))
    billingCycle = "February (2024)"

    resourcegroups.step_one(billing, chargeback)
    rg.rgComparer(chargeback)

    resourcegroups.step_two(chargeback, chargeback2)
    summary2.step_three(chargeback, billingCycle)
    summary2.step_four(chargeback, chargeback2)

    wb = load_workbook(chargeback)
    summarySheet = pd.read_excel(chargeback, sheet_name = "Summary")
    summary.createChargebacks(wb, chargeback, summarySheet, 0.0305)